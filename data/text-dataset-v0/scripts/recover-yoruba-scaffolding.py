#!/usr/bin/env python3
"""Recover reviewed Yoruba scaffolding from the linguist's Unicode XLSX."""

import csv
import os
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DRAFT = ROOT / "data/scaffolding/yo-en.DRAFT.csv"
WORKBOOK = ROOT / "data/scaffolding/yo-en.DRAFT.20260813150538134.xlsx"
OUTPUT = ROOT / "data/scaffolding/yo-en.csv"
REVIEW_COLUMNS = ("verdict", "corrected_text")


def text(value):
    return "" if value is None else str(value)


with DRAFT.open(encoding="utf-8-sig", newline="") as source:
    reader = csv.DictReader(source)
    header = reader.fieldnames
    draft_rows = list(reader)

if not header or any(column not in header for column in REVIEW_COLUMNS):
    raise SystemExit("Draft CSV does not have the expected review schema.")

workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
sheet = workbook.active
values = list(sheet.iter_rows(values_only=True))
workbook_header = [text(value) for value in values[0]]
if workbook_header != header:
    raise SystemExit("Workbook and draft headers differ; refusing an unsafe merge.")

workbook_rows = [dict(zip(header, map(text, row))) for row in values[1:]]
if len(workbook_rows) != len(draft_rows):
    raise SystemExit("Workbook and draft row counts differ; refusing an unsafe merge.")

base_columns = [column for column in header if column not in REVIEW_COLUMNS]
for position, (draft, reviewed) in enumerate(zip(draft_rows, workbook_rows), start=2):
    if draft["act_id"] != reviewed["act_id"]:
        raise SystemExit(f"Row {position} act_id mismatch; refusing an unsafe merge.")
    for column in base_columns:
        if draft[column] != reviewed[column]:
            raise SystemExit(
                f"Row {position} differs in base column {column}; refusing an unsafe merge."
            )
    for column in REVIEW_COLUMNS:
        draft[column] = reviewed[column]

if any(not row["verdict"].strip() for row in draft_rows):
    raise SystemExit("The workbook has blank verdicts; refusing to publish an incomplete merge.")

temporary = OUTPUT.with_suffix(".csv.tmp")
with temporary.open("w", encoding="utf-8-sig", newline="") as target:
    writer = csv.DictWriter(target, fieldnames=header, lineterminator="\n")
    writer.writeheader()
    writer.writerows(draft_rows)
os.replace(temporary, OUTPUT)

print(
    f"Recovered {len(draft_rows)} Yoruba rows with "
    f"{sum(bool(row['corrected_text'].strip()) for row in draft_rows)} corrections."
)
